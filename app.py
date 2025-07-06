import streamlit as st
import pymysql
import openpyxl 
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import time


conexao = pymysql.connect(
    host='localhost',
    user='root',
    password='',
    database='rosa'
)

cursor = conexao.cursor()

data_min = datetime(1900, 1, 1)
data_max = datetime.now()

def main():
    st.set_page_config(page_title='Outubro Rosa', page_icon='icone.jpg')
    st.title("Formulário Outubro Rosa!")
    st.image('senac_outubro_rosa_.jpg', width=700) 

    form = st.form(key = "Clientes", clear_on_submit=True)
    with form:
        nome = st.text_input('NOME:')
        curso = st.selectbox('CURSOS:', ('Recurso Humanos', 'Secretariado', 'Técnico de ADM', 'Costura', 'Cabelereira', 'Programador de Sistemas (o melhor que tem)', 'Segurança do Trabalho', 'Coordenação'),
        index = None,
        placeholder = 'Informe seu Curso')
        data_nascimento = st.date_input('DATA DE NASCIMENTO:',
            min_value=data_min,
            value=None)
        st.write('M = Masculino = F = Feminino = O = Other')
        sexo = st.selectbox('SEXO:', ['M', 'F', 'O'])

        question1 = st.radio('Você costuma realizar autoexames das mamas? Se sim, com que frequência?', ('Mensalmente', 'A cada três meses', 'Raramente', 'Nunca'),
            index = None)

        question2 = st.radio('Você está familiarizada com os fatores de risco relacionados ao câncer de mama?',('Sim', 'Não'),
            index = None)

        question3 = st.radio('Qual é sua principal fonte de informação sobre o câncer de mama?',('Internet', 'Médicos/Profissionais de saúde', 'Amigos/familiares', 'Materiais impressos(Panfletos, cartazes)'),
            index = None)

        question4 = st.radio('Na sua opinião, o que é mais importante para a prevenção do câncer de mama?',('Exames regulares', 'Manter um estilo de vida saudável', 'Ter apoio emocional e psicológico'),
            index = None)
  
        botao = form.form_submit_button("Enviar")

    if botao:
   
            if nome and curso and data_nascimento and sexo and question1 and question2 and question3 and question4 !='':
                comando = f'INSERT INTO dados (yname, course, gender, date_of_birth, question1, question2, question3, question4) VALUES ("{nome}", "{curso}", "{sexo}", "{data_nascimento}", "{question1}", "{question2}", "{question3}", "{question4}")'
                cursor.execute(comando)
                conexao.commit()
                sucesso = st.success('Obrigado pela Presença! Ame-se Realize o Auto Exame')
                st.balloons()
                time.sleep(5)
                sucesso.empty()

            else:
                st.error("Por favor, insira dados válidos")


    if nome == 'Final':
        comando2 = f'select * from dados'
        cursor.execute(comando2)
        resultado = cursor.fetchall()
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active


        for col_num, column_name in enumerate([desc[0] for desc in cursor.description], start=1):
            sheet.cell(row=1, column= col_num, value= column_name)

        row_num = 2
        for registro in resultado:
            for col_num, cell_value in enumerate(registro, start=1):
                sheet.cell(row= row_num, column= col_num, value= cell_value)
            row_num +=1

        workbook.save('dados.xlsx')

        st.title('Gráfico')

        df = pd.read_excel('dados.xlsx')
        
        st.write('Dados do arquivo:')
        st.dataframe(df)

        perguntas = {
            'question1': 'Autoexame das mamas',
            'question2': 'Conhecimento sobre fatores de risco',
            'question3': 'Fonte de informação sobre o câncer de mama',
            'question4': 'Opinião sobre prevenção'
        }

        pergunta_escolhida = st.selectbox('Escolha a pergunta:', list(perguntas.keys()), format_func=lambda x: perguntas[x])

        agrupado = df.groupby(['gender', pergunta_escolhida]).size().unstack(fill_value=0)
        agrupado = agrupado.rename(index={'M': 'Homens', 'F': 'Mulheres', 'O': 'Outros'})

        fig, ax = plt.subplots(figsize=(10, 5))
        agrupado.plot(kind='bar', stacked=True, ax=ax, colormap='Pastel1')
        ax.set_title(f'Distribuição das respostas - {perguntas[pergunta_escolhida]}')
        ax.set_ylabel('Número de pessoas')
        ax.set_xlabel('Gênero')
        ax.legend(title='Resposta', bbox_to_anchor=(1.05, 1), loc='upper left')

        st.pyplot(fig)
 
if __name__ == '__main__':
    main()
